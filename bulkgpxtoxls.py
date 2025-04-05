import tkinter as tk
from tkinter import filedialog
import os
import gpxpy
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def process_gpx_file(file_path):
    skipped_files = []
    try:
        gpx = gpxpy.parse(open(file_path, 'r'))
        wb = Workbook()
        ws = wb.active
        ws.title = "GPX Data"
        headers = ["type", "time", "latitude", "longitude", "altitude (m)"]
        for i, header in enumerate(headers, start=1):
            ws[f"{get_column_letter(i)}1"] = header

        row = 2
        for track in gpx.tracks:
            for segment in track.segments:
                for point in segment.points:
                    try:
                        ws[f"A{row}"] = "T"
                        ws[f"B{row}"] = point.time.strftime("%d/%m/%Y %H:%M:%S")
                        ws[f"C{row}"] = str(point.latitude).replace(".", ",")
                        ws[f"D{row}"] = str(point.longitude).replace(".", ",")
                        ws[f"E{row}"] = str(point.elevation).replace(".", ",")
                        row += 1
                    except Exception as e:
                        print(f"Error processing point: {e}")
        output_folder = r"C:\Users\Usuario\Desktop\Geocode"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        output_file = os.path.join(output_folder, os.path.basename(file_path).replace(".gpx", ".xlsx"))
        wb.save(output_file)
        return skipped_files
    except Exception as e:
        skipped_files.append(file_path)
        print(f"Error processing file: {e}")
        return skipped_files

def main():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_paths = filedialog.askopenfilenames(title="Select GPX Files", filetypes=[("GPX files", "*.gpx")])
    skipped_files = []
    for file_path in file_paths:
        skipped_files.extend(process_gpx_file(file_path))

    if skipped_files:
        skipped_files_message = "\n".join(skipped_files)
        skipped_files_message = f"The following files were skipped due to errors:\n{skipped_files_message}"
    else:
        skipped_files_message = "All files processed successfully."
    completed_message = f"Process completed.\n{skipped_files_message}"
    tk.messagebox.showinfo("Process Completed", completed_message)

if __name__ == "__main__":
    main()
